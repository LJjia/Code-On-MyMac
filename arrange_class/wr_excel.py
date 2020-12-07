#!/usr/bin/env python
# -*- coding:utf-8 _*-
__author__ = 'LJjia'
# *******************************************************************
#     Filename @  wr_excel.py
#       Author @  Jia Liangjun
#  Create date @  2020/08/30 18:24
#        Email @  LJjiahf@163.com
#  Description @  跟排课程序对接的读取excel接口
# ********************************************************************

import data_sheet
import openpyxl
import time

wb_src = openpyxl.load_workbook('排课参数表.xlsx')

school_shcedule_tab = wb_src.get_sheet_by_name('学校日程表')
school_week_shcedule = [[], [], [], [], [], [], []]
if_line_head = 0
for day, col in enumerate(school_shcedule_tab.columns):
    if if_line_head == 0:
        if_line_head += 1
        continue
    if_day = 0
    for cell in col:
        if if_day == 0:
            if_day += 1
            continue
        if cell.value == '上课':
            school_week_shcedule[day - 1].append(1)
        else:
            school_week_shcedule[day - 1].append(0)
print('学校日程表', school_week_shcedule)

teacher_course_tab = wb_src.get_sheet_by_name('教师课时统计')
teacher_course_dict = {}
teacher_list = []
if_line_head = 0
for day, row in enumerate(teacher_course_tab.rows):
    if if_line_head == 0:
        if_line_head += 1
        continue
    teacher_list.append(row[0].value)
    teacher_course_dict[row[0].value] = int(row[1].value)
print('教师上课字典', teacher_course_dict)
print('教师名', teacher_list)

class_num_tab = wb_src.get_sheet_by_name('班级数量')
class_num = class_num_tab.cell(1, 1).value
print('班级数量', class_num)


range_obj = data_sheet.Course(school_schedule=school_week_shcedule, teacher_list=teacher_list,
                              teacher_class_num_dict=teacher_course_dict,
                              class_num=class_num)
all_class_schedule = range_obj.arrange_class()


def write_excel_head(ws,ws_src):
    '''
    写班级excel排序表的表头
    :param ws: excel表对象
    :return:
    '''

    for period, row in enumerate(ws_src.rows):
        for day,cell in enumerate(row):
            ws.cell(row=period+1,column=day+1,value=cell.value)


wb_dst = openpyxl.Workbook()
del_tab = wb_dst.active
wb_dst.remove(del_tab)
for cnt in range(class_num):
    ws = wb_dst.create_sheet("班级%s" % (cnt+1))
    class_schedule = all_class_schedule[cnt]
    write_excel_head(ws,school_shcedule_tab)
    for day in range(len(class_schedule)):
        for period in range(len(class_schedule[day])):
            if(class_schedule[day][period]==0):
                pass
                #ws.cell(row=2 + period, column=2 + day, value='')
            else:
                ws.cell(row=2+period,column=2+day,value=class_schedule[day][period])
wb_dst.save('班级课表.xlsx')
wb_src.close()
time.sleep(2)

